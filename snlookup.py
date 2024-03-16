import openpyxl
from openpyxl import Workbook
import requests
from bs4 import BeautifulSoup
import re
 
wb = openpyxl.load_workbook('/Users/brucnoa2185/Documents/TestSpreadsheet.xlsx')
sheet = wb['Sheet1'] 

new_wb = Workbook()
new_sheet = new_wb.active
new_sheet.append(['Serial Number', 'Part Number', 'Model Name'])  
 
base_url = "https://partsurfer.hp.com/partsurfer/?searchtext={}&searchby=swp"
counter = 0
for row in sheet.iter_rows(min_row=2, values_only=True):
    if counter >= 10:
        break
    
    serial_number = row[7]
    print(f"Processing serial number: {serial_number}")
    url = base_url.format(serial_number)
    print(url)
    response = requests.get(url)
    soup = BeautifulSoup(response.text, 'html.parser')
    

    #product_number_element = soup.find_all(string=re.compile("HP"))
    product_number_element = soup.find(string="Product Number :  ")
    #product_number_element = soup.find_all(text=re.compile("HP"))
    print(product_number_element)
    if product_number_element:
        product_number = product_number_element.find_next().text
    else:
        product_number = "Not Found"
 
    description_element = soup.find(string="Description :  ")
    print(description_element)
    if description_element:
        description = description_element.find_next().text
    else:
        description = "Not Found"
    
    new_sheet.append([serial_number, product_number, description])
    counter += 1
 


new_wb.save('/Users/brucnoa2185/Documents/output.xlsx')
